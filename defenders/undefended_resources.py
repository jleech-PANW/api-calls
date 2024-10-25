import requests, json, configparser, openpyxl
from datetime import date
from openpyxl import load_workbook

today = date.today()

def auth_func():
    config = configparser.ConfigParser()
    config.read('config.ini')
    login_url = config.get('prismacloud', 'cspm_api_url')
    username = config.get('prismacloud', 'username')
    password = config.get('prismacloud', 'password')
    try:
        url = f'{login_url}/login'
        payload = f"{{\n  \"password\": \"{password}\",\n  \"username\": \"{username}\"\n}}"
        headers = {
            'Content-Type': 'application/json; charset=UTF-8',
            'Accept': 'application/json; charset=UTF-8'
        }
        response = requests.request("POST", url, headers=headers, data=payload)
        response_json = response.json()
        if 'token' in response_json:
            print("Retrieved Token")
            return response_json['token']
        else:
            print("Failed to get authentication token")
            return None 
    except Exception as e:
        print(f"Failed to get authentication token: {e}")
        return None    


def entity_report():
  print(f"Creating {provider} Entity Report")
  token = auth_func()
  config = configparser.ConfigParser()
  config.read('config.ini')
  base_url = config.get('prismacloud', 'cwp_api_url')
  headers = {
      'Content-Type': 'application/json',
      'x-redlock-auth': token
  }
  payload = {}  
  file_name = f"undefended_resources_{today}_{provider}.xlsx"
  url = f'{base_url}/api/v33.01/cloud/discovery/entities?defended=false&offset=0&provider={provider}'
  response = requests.request("GET", url, headers=headers, data=payload)
  print(f"Initial response Code: {response.status_code}")
  try:
    data = json.loads(response.text)
  except json.JSONDecodeError:
     print(f"No {provider} entities.")
     return
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = "Undefended Entities"
  wb_headers = ["Name", "ARN", "AccountID", "Region", "Workload Type"]
  ws.append(wb_headers)
  offset = 0
  if data is None:
     print(f"No {provider} entities.")
     return
  print(f"Creating {file_name}")
  for item in data:
    ws.append([item.get('name', ''), item.get('arn', ''), item.get('accountID', ''), item.get('region', ''), item.get('serviceType', '')])
  
  while data is not None:
    offset += 50
    url = f'{base_url}/api/v33.01/cloud/discovery/entities?defended=false&offset={offset}&provider={provider}'
    response = requests.request("GET", url, headers=headers, data=payload)
    try:
      data = json.loads(response.text)
    except json.JSONDecodeError:
       wb.save(file_name)
       print(f"{provider} Entity Run Complete")
       return
    if data is None:
      wb.save(file_name)
      print(f"{provider} Entity Run Complete")
      return
    for item in data:
      ws.append([item.get('name', ''), item.get('arn', ''), item.get('accountID', ''), item.get('region', ''), item.get('serviceType', '')])
    wb.save(file_name)
    print(f"Saved up to {offset}. Response Code: {response.status_code} Provider: {provider}")
    if offset % 1000 == 0: 
      token = auth_func()

def vm_report():
  print(f"Creating {provider} VM Report")
  token = auth_func()
  config = configparser.ConfigParser()
  config.read('config.ini')
  base_url = config.get('prismacloud', 'cwp_api_url')
  headers = {
      'Content-Type': 'application/json',
      'x-redlock-auth': token
  }
  payload = {}  
  file_name = f"undefended_resources_{today}_{provider}.xlsx"
  url = f'{base_url}/api/v33.01/cloud/discovery/vms?offset=0&hasDefender=false&provider={provider}'
  response = requests.request("GET", url, headers=headers, data=payload)
  print(f"Initial response Code: {response.status_code}")
  try:
    data = json.loads(response.text)
  except json.JSONDecodeError:
     print(f"No {provider} VMs.")
     return
  try:
    wb = load_workbook(file_name)
  except FileNotFoundError:
     wb = openpyxl.Workbook()
  ws2 = wb.create_sheet(title="Undefended VMs")
  wb_headers = ["Name", "ARN", "AccountID", "Region", "Hostname"]
  ws2.append(wb_headers)
  offset = 0
  if data is None:
    print(f"No {provider} VMs.")
    return
  print(f"Creating {file_name}")
  for item in data:
    ws2.append([item.get('name', ''), item.get('arn', ''), item.get('accountID', ''), item.get('region', ''), item.get('hostname', '')])
  
  while data is not None:
    offset += 50
    url = f'{base_url}/api/v33.01/cloud/discovery/vms?offset={offset}&hasDefender=false&provider={provider}'
    response = requests.request("GET", url, headers=headers, data=payload)
    try:
      data = json.loads(response.text)
    except json.JSONDecodeError:
       wb.save(file_name)
       print(f"{provider} VM Run Complete")
       return
    if data is None:
      wb.save(file_name)
      print(f"{provider} VM Run Complete")
      return
    for item in data:
      ws2.append([item.get('name', ''), item.get('arn', ''), item.get('accountID', ''), item.get('region', ''), item.get('hostname', '')])
    wb.save(file_name)
    print(f"Saved up to {offset}. Response Code: {response.status_code} Provider: {provider}")
    if offset % 1000 == 0: 
      token = auth_func()



providers = ["others", "alibaba", "oci", "gcp", "azure", "aws" ]
for provider in providers:
  entity_report()
  vm_report()

