import configparser
import pandas as pd
from pcpi import saas_session_manager
from openpyxl import load_workbook

def initialize_environment():
    dbconfig = configparser.ConfigParser()
    dbconfig.read('config-ot.ini')
    config = {
        'CSPM_ENDPOINT': dbconfig.get('prismacloud', 'cspm_api_url'),
        'USERNAME': dbconfig.get('prismacloud', 'username'),
        'PASSWORD': dbconfig.get('prismacloud', 'password'),
    }
    return config

def initialize_services(config):
    session_manager = saas_session_manager.SaaSSessionManager(
        'Tenant', config['USERNAME'], config['PASSWORD'], config['CSPM_ENDPOINT']
    )
    services = {
        'cspm': session_manager.create_cspm_session(),
        'cwp': session_manager.create_cwp_session(),
        'container_client': None
    }
    return services

def versions_check():
    config = initialize_environment()
    services = initialize_services(config)
    cspm = services['cspm']
    cwp = services['cwp']

    response = cwp.request('GET', f'api/v1/version')
    global current_major, last_major
    current_release = response.text.replace('"', '')
    current_major = int(current_release.split(".")[0])
    last_major = int(current_release.split(".")[0]) - 1
#    print(current_major, last_major)
    return 

def currentmajor_defenders():
    #setup
    print("Getting Current Release Defenders")
    config = initialize_environment()
    services = initialize_services(config)
    cspm = services['cspm']
    cwp = services['cwp']
    offset = 0
    #First 50 Defenders
    response = cwp.request('GET', f'/api/v33.03/defenders?limit=50')
    defenders = pd.DataFrame(response.json())[['hostname', 'version', 'cluster', 'lastModified', 'category', 'cloudMetadata']]
    defenders['provider'] = defenders['cloudMetadata'].apply(lambda x: x.get('provider', 'N/A'))
    defenders['accountID'] = defenders['cloudMetadata'].apply(lambda x: x.get('accountID', 'N/A'))
    defenders = defenders.drop(columns=['cloudMetadata'])
    defenders = defenders[defenders['version'].notna()]
    defenders = defenders[defenders['version'] != '']
    defenders['major_version'] = defenders['version'].str.split(".").str[0].astype(int)
    defenders = defenders[defenders['major_version'] == current_major]
    defenders = defenders.drop(columns=['major_version'])
    defenders.to_excel("currentmajor_defenders.xlsx", index=False, sheet_name="Current Release Defenders")

    #offsets
    while response.json() is not None:
        offset += 50
        response = cwp.request('GET', f'/api/v33.03/defenders?limit=50&offset={offset}')
        if response.json() is not None:
            defenders = pd.DataFrame(response.json())[['hostname', 'version', 'cluster', 'lastModified', 'category', 'cloudMetadata']]
            defenders['provider'] = defenders['cloudMetadata'].apply(lambda x: x.get('provider', 'N/A'))
            defenders['accountID'] = defenders['cloudMetadata'].apply(lambda x: x.get('accountID', 'N/A'))
            defenders = defenders.drop(columns=['cloudMetadata'])
            defenders = defenders[defenders['version'].notna()]
            defenders = defenders[defenders['version'] != '']
            defenders['major_version'] = defenders['version'].str.split(".").str[0].astype(int)
            defenders = defenders[defenders['major_version'] == current_major]
            defenders = defenders.drop(columns=['major_version'])
            with pd.ExcelWriter("currentmajor_defenders.xlsx", mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
                book = load_workbook("currentmajor_defenders.xlsx")
                sheet = book["Current Release Defenders"]
                last_row = sheet.max_row
                defenders.to_excel(writer, sheet_name="Current Release Defenders", index=False, header=False, startrow=last_row)
        else:
            return
        
def lastmajor_defenders():
    print("Getting Last Major Release Defenders")
    #setup
    config = initialize_environment()
    services = initialize_services(config)
    cspm = services['cspm']
    cwp = services['cwp']
    offset = 0
    #First 50 Defenders
    response = cwp.request('GET', f'/api/v33.03/defenders?limit=50')
    defenders = pd.DataFrame(response.json())[['hostname', 'version', 'cluster', 'lastModified', 'category', 'cloudMetadata']]
    defenders['provider'] = defenders['cloudMetadata'].apply(lambda x: x.get('provider', 'N/A'))
    defenders['accountID'] = defenders['cloudMetadata'].apply(lambda x: x.get('accountID', 'N/A'))
    defenders = defenders.drop(columns=['cloudMetadata'])
    defenders = defenders[defenders['version'].notna()]
    defenders = defenders[defenders['version'] != '']
    defenders['major_version'] = defenders['version'].str.split(".").str[0].astype(int)
    defenders = defenders[defenders['major_version'] == last_major]
    defenders = defenders.drop(columns=['major_version'])
    defenders.to_excel("lastmajor_defenders.xlsx", index=False, sheet_name="Last Release Defenders")

    #offsets
    while response.json() is not None:
        offset += 50
        response = cwp.request('GET', f'/api/v33.03/defenders?limit=50&offset={offset}')
        if response.json() is not None:
            defenders = pd.DataFrame(response.json())[['hostname', 'version', 'cluster', 'lastModified', 'category', 'cloudMetadata']]
            defenders['provider'] = defenders['cloudMetadata'].apply(lambda x: x.get('provider', 'N/A'))
            defenders['accountID'] = defenders['cloudMetadata'].apply(lambda x: x.get('accountID', 'N/A'))
            defenders = defenders.drop(columns=['cloudMetadata'])
            defenders = defenders[defenders['version'].notna()]
            defenders = defenders[defenders['version'] != '']
            defenders['major_version'] = defenders['version'].str.split(".").str[0].astype(int)
            defenders = defenders[defenders['major_version'] == last_major]
            defenders = defenders.drop(columns=['major_version'])
            with pd.ExcelWriter("lastmajor_defenders.xlsx", mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
                book = load_workbook("lastmajor_defenders.xlsx")
                sheet = book["Last Release Defenders"]
                last_row = sheet.max_row
                defenders.to_excel(writer, sheet_name="Last Release Defenders", index=False, header=False, startrow=last_row)
        else:
            return

def outdated_defenders():
    print("Getting Outdated Defenders")
    #setup
    config = initialize_environment()
    services = initialize_services(config)
    cspm = services['cspm']
    cwp = services['cwp']
    offset = 0
    #First 50 Defenders
    response = cwp.request('GET', f'/api/v33.03/defenders?limit=50')
    defenders = pd.DataFrame(response.json())[['hostname', 'version', 'cluster', 'lastModified', 'category', 'cloudMetadata']]
    defenders['provider'] = defenders['cloudMetadata'].apply(lambda x: x.get('provider', 'N/A'))
    defenders['accountID'] = defenders['cloudMetadata'].apply(lambda x: x.get('accountID', 'N/A'))
    defenders = defenders.drop(columns=['cloudMetadata'])
    defenders['version'] = defenders['version'].replace(['', None], '0.0.0.0')
    defenders['major_version'] = defenders['version'].str.split(".").str[0].astype(int)
    defenders = defenders[defenders['major_version'] < last_major]
    defenders = defenders.drop(columns=['major_version'])
    defenders.to_excel("outdated_defenders.xlsx", index=False, sheet_name="Out Dated Defenders")

    #offsets
    while response.json() is not None:
        offset += 50
        response = cwp.request('GET', f'/api/v33.03/defenders?limit=50&offset={offset}')
        if response.json() is not None:
            defenders = pd.DataFrame(response.json())[['hostname', 'version', 'cluster', 'lastModified', 'category', 'cloudMetadata']]
            defenders['provider'] = defenders['cloudMetadata'].apply(lambda x: x.get('provider', 'N/A'))
            defenders['accountID'] = defenders['cloudMetadata'].apply(lambda x: x.get('accountID', 'N/A'))
            defenders = defenders.drop(columns=['cloudMetadata'])
            defenders['version'] = defenders['version'].replace(['', None], '0.0.0.0')
            defenders['major_version'] = defenders['version'].str.split(".").str[0].astype(int)
            defenders = defenders[defenders['major_version'] < last_major]
            defenders = defenders.drop(columns=['major_version'])
            with pd.ExcelWriter("outdated_defenders.xlsx", mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
                book = load_workbook("outdated_defenders.xlsx")
                sheet = book["Out Dated Defenders"]
                last_row = sheet.max_row
                defenders.to_excel(writer, sheet_name="Out Dated Defenders", index=False, header=False, startrow=last_row)
        else:
            return

versions_check()
currentmajor_defenders()
lastmajor_defenders()
outdated_defenders()
