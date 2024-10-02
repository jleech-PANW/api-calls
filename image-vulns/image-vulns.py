import requests, json, configparser, openpyxl
from openpyxl import load_workbook

def auth_func():
    config = configparser.ConfigParser()
    config.read('config.ini')
    login_url = config.get('prismacloud', 'cspm_api_url')
    username = config.get('prismacloud', 'username')
    password = config.get('prismacloud', 'password')
    print("Logging In...")
    try:
        url = f'{login_url}/login'
        payload = f"{{\n  \"password\": \"{password}\",\n  \"username\": \"{username}\"\n}}"
        headers = {
            'Content-Type': 'application/json; charset=UTF-8',
            'Accept': 'application/json; charset=UTF-8'
        }
        response = requests.request("POST", url, headers=headers, data=payload)
        response_json = response.json()
        if 'token' in response_json:
            print("Login Successful")
            return response_json['token']
        else:
            print("Failed to get authentication token")
            return None 

    except Exception as e:
        print(f"Failed to get authentication token: {e}")
        return None    
token = auth_func()
def get_image_vulns():
    print("Retrieving Image Vulnerabilities from Prisma...")
    config = configparser.ConfigParser()
    config.read('config.ini')
    base_url = config.get('prismacloud', 'cwp_api_url')
    try:
        headers = {
            'Content-Type': 'application/json',
            'x-redlock-auth': f'{token}'
        }
        payload = {}
        url = f'{base_url}/api/v33.00/images'
        
        #Troubleshooting print
        #print(f"sending request to {url} with headers: \n {headers}")
        
        response = requests.request("GET", url, headers=headers, data=payload)
        data = json.loads(response.text)
        filtered_data = [{k: v for k, v in entry.items() if k in ["repoTag", "instances", "vulnerabilities"]} for entry in data]
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        fieldnames = ['CVE', 'image']
        sheet.append(fieldnames)
        for entry in filtered_data:
            for instance in entry['instances']:
                image = instance['image']
                for vuln in entry['vulnerabilities']:
                    cve = vuln['cve']
                
                    sheet.append([cve, image])
        workbook.save('image-report.xlsx')
        
        print("Successful")
        return None
    except Exception as e:
        print(f"Failed to get image vulns from prisma: {e}")
        return None 


def get_container_count():
    print("Getting vulnerable container count...")
    config = configparser.ConfigParser()
    config.read('config.ini')
    base_url = config.get('prismacloud', 'cwp_api_url')    
    try:
        headers = {
            'Content-Type': 'application/json',
            'x-redlock-auth': f'{token}'
        }
        payload = {}
        workbook = load_workbook('image-report.xlsx')
        sheet = workbook.active
        image_column = "B"
        response_column = "C"
        processed_images = {}
        for row in range(2, sheet.max_row + 1):
            image_name = sheet[f'{image_column}{row}'].value
            if image_name:
                if image_name not in processed_images:
                    url = f'{base_url}/api/v33.00/containers/count?image={image_name}'
                    response = requests.request("GET", url, headers=headers, data=payload)
                    if response.status_code == 200:
                        processed_images[image_name] = str(response.json())
                    else:
                        processed_images[image_name] = f"Failed: {response.status_code}"
            sheet[f'{response_column}{row}'] = processed_images[image_name]  
        workbook.save('image-report.xlsx')
        print("Successful")
    except Exception as e:
        print(f"Failed to get image vulns from prisma: {e}")
        return None 

get_image_vulns()
get_container_count()
